import openpyxl
print(openpyxl.__version__)

import os
os.chdir('C:\\Users\Administrator\Desktop')
wb = openpyxl.load_workbook('Practice.xlsx')
print (type(wb))
sheet = wb['Sheet1']
print (sheet['A1'].value)


# Real world and excel, numbers start with 1, so it is needed to do so
for i in range (1,sheet.max_row):
    #to manually manipulate the the date of a excel sheet for all rows
    X = sheet.cell(row=i, column=1).value
    theDate = str(X)
    print (theDate)
    Year = theDate[0:4]
    Month= theDate[5:7]
    Day = theDate[8:10]

    newDate = Month + '/' + Day + '/' + Year
    print (theDate)
    print(newDate)
    #to save 
    wb.save('Practice.xlsx')
    print("It is saved")
